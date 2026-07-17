"""Adaptateur fichiers CSV / Excel (V1.0).

Les fichiers plats deviennent une vraie source SQL : Noreon les matérialise
dans une base **SQLite** locale (un fichier CSV → une table ; un classeur
Excel → une table par feuille), avec inférence de type. Le scanner, le
profileur et le chat fonctionnent alors à l'identique des bases relationnelles.

Lecture seule par construction : les requêtes ouvrent SQLite en mode `ro`.
"""
from __future__ import annotations

import csv
import os
import re
import sqlite3
from contextlib import contextmanager
from typing import Iterator

from app.core.logging import get_logger
from app.services.sources.base import (
    ColumnInfo,
    ScanResult,
    SourceAdapter,
    TableInfo,
    infer_relations,
)

log = get_logger("noreon.source.files")

_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+[.,]?\d*$")


def _sanitize(name: str) -> str:
    base = re.sub(r"[^0-9a-zA-Z_]+", "_", os.path.splitext(os.path.basename(name))[0]).strip("_")
    if not base or base[0].isdigit():
        base = "t_" + base
    return base.lower()


def _infer_type(values: list[str]) -> str:
    non_empty = [v for v in values if v not in (None, "")]
    if not non_empty:
        return "TEXT"
    if all(_INT_RE.match(str(v)) for v in non_empty):
        return "INTEGER"
    if all(_FLOAT_RE.match(str(v).replace(",", ".")) for v in non_empty):
        return "REAL"
    return "TEXT"


def _convert(value: str, sqlite_type: str):
    if value in (None, ""):
        return None
    if sqlite_type == "INTEGER":
        try:
            return int(value)
        except ValueError:
            return value
    if sqlite_type == "REAL":
        try:
            return float(str(value).replace(",", "."))
        except ValueError:
            return value
    return value


class FileAdapter(SourceAdapter):
    engine = "csv"
    dialect = "sqlite"

    def __init__(self, config) -> None:
        super().__init__(config)
        self.engine = config.engine  # csv | excel
        self._sqlite = self._sqlite_path()

    def _sqlite_path(self) -> str:
        from app.core.config import settings

        explicit = self.config.options.get("sqlite_path")
        if explicit:
            return explicit
        os.makedirs(settings.data_dir, exist_ok=True)
        base = _sanitize(self.config.file_path or "source")
        return os.path.join(settings.data_dir, f"{base}.noreon.sqlite")

    # --- matérialisation ---
    def _needs_build(self) -> bool:
        if not os.path.exists(self._sqlite):
            return True
        src = self.config.file_path
        return bool(src and os.path.exists(src) and os.path.getmtime(src) > os.path.getmtime(self._sqlite))

    def _ensure_materialized(self) -> None:
        if not self._needs_build():
            return
        src = self.config.file_path
        if not src or not os.path.exists(src):
            raise FileNotFoundError(f"Fichier source introuvable : {src}")
        tmp = self._sqlite + ".building"
        if os.path.exists(tmp):
            os.remove(tmp)
        con = sqlite3.connect(tmp)
        try:
            if (self.config.options.get("format") or self.engine) == "excel":
                self._load_excel(con, src)
            else:
                self._load_csv(con, src)
            con.commit()
        finally:
            con.close()
        os.replace(tmp, self._sqlite)

    def _create_and_fill(self, con: sqlite3.Connection, table: str, headers: list[str], rows: list[list]) -> None:
        headers = [h.strip() or f"col_{i}" for i, h in enumerate(headers)]
        types = [
            _infer_type([r[i] if i < len(r) else "" for r in rows[:500]])
            for i in range(len(headers))
        ]
        cols_ddl = ", ".join(f'"{h}" {t}' for h, t in zip(headers, types))
        con.execute(f'CREATE TABLE "{table}" ({cols_ddl})')
        placeholders = ", ".join("?" * len(headers))
        con.executemany(
            f'INSERT INTO "{table}" VALUES ({placeholders})',
            [[_convert(r[i] if i < len(r) else None, types[i]) for i in range(len(headers))] for r in rows],
        )

    def _load_csv(self, con: sqlite3.Connection, path: str) -> None:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(fh, dialect)
            rows = list(reader)
        if not rows:
            raise ValueError("Fichier CSV vide.")
        headers, data = rows[0], rows[1:]
        # Nom de table = nom d'origine du fichier (sans le préfixe technique d'upload).
        origin = self.config.options.get("original_name") or path
        self._create_and_fill(con, _sanitize(origin), headers, data)

    def _load_excel(self, con: sqlite3.Connection, path: str) -> None:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        made = 0
        for ws in wb.worksheets:
            rows = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
            rows = [r for r in rows if any(str(c).strip() for c in r)]
            if len(rows) < 1:
                continue
            headers = [str(h) for h in rows[0]]
            self._create_and_fill(con, _sanitize(ws.title), headers, rows[1:])
            made += 1
        if made == 0:
            raise ValueError("Classeur Excel sans feuille exploitable.")

    @contextmanager
    def _open(self) -> Iterator[sqlite3.Connection]:
        self._ensure_materialized()
        con = sqlite3.connect(f"file:{self._sqlite}?mode=ro", uri=True)
        try:
            yield con
        finally:
            con.close()

    # --- primitives ---
    def fetch(self, sql: str, params: tuple | None = None) -> tuple[list[str], list[list]]:
        with self._open() as con:
            cur = con.execute(sql, params or ())
            cols = [d[0] for d in cur.description] if cur.description else []
            rows = [list(r) for r in cur.fetchall()]
        return cols, rows

    def qualified(self, schema: str, table: str) -> str:
        return self.quote_ident(table)  # SQLite : pas de schéma

    # --- connexion / conformité ---
    def test_connection(self) -> dict:
        try:
            self._ensure_materialized()
            _, rows = self.fetch("SELECT sqlite_version()")
            fmt = self.config.options.get("format") or self.engine
            return {"ok": True, "server_version": f"Fichier {fmt.upper()} (SQLite {rows[0][0]})", "error": None}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "server_version": None, "error": str(exc)}

    def check_read_only(self) -> dict:
        # Source fichier : lecture seule par construction (ouverture SQLite en mode ro).
        return {"read_only": True, "detail": "Source fichier — lecture seule par construction.", "error": None}

    # --- introspection ---
    def introspect(self) -> ScanResult:
        with self._open() as con:
            tnames = [r[0] for r in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall()]
            tables: dict[tuple[str, str], TableInfo] = {}
            for name in tnames:
                info = con.execute(f'PRAGMA table_info("{name}")').fetchall()
                count = con.execute(f'SELECT count(*) FROM "{name}"').fetchone()[0]
                ti = TableInfo(schema="main", name=name, table_type="table",
                               estimated_rows=int(count), comment=None)
                for cid, cname, ctype, notnull, default, pk in info:
                    ti.columns.append(ColumnInfo(
                        name=cname, ordinal=int(cid) + 1, data_type=(ctype or "TEXT").lower(),
                        is_nullable=not bool(notnull), default=default, is_pk=bool(pk),
                    ))
                tables[("main", name)] = ti
        # Fichiers : pas de FK déclarées → uniquement relations inférées.
        relations = infer_relations(list(tables.values()), set())
        return ScanResult(tables=list(tables.values()), relations=relations)

    # --- exécution ---
    def _execute(self, sql: str, timeout_seconds: int) -> tuple[list[str], list[list]]:
        return self.fetch(sql)
